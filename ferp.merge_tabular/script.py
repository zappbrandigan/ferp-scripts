from __future__ import annotations

import csv
import hashlib
from collections import defaultdict
from pathlib import Path
from typing import Iterable, Iterator, List

import openpyxl

from ferp.fscp.scripts import sdk

SUPPORTED_EXTS = {".csv", ".xlsx"}


def _read_header(path: Path) -> List[str]:
    if path.suffix.lower() == ".csv":
        with path.open("r", newline="", encoding="utf-8") as fh:
            reader = csv.reader(fh)
            try:
                header = next(reader)
            except StopIteration as exc:  # noqa: B904
                raise ValueError("File is empty") from exc
            return header
    return _read_xlsx_header(path)


def _read_xlsx_header(path: Path) -> List[str]:
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.active
    if sheet is None:
        raise ValueError("Workbook has no active sheet")
    header_row = next(sheet.iter_rows(max_row=1, values_only=True), None)
    if header_row is None:
        raise ValueError("File is empty")
    return ["" if cell is None else str(cell) for cell in header_row]


def _iter_rows(path: Path) -> Iterator[Iterable[object]]:
    if path.suffix.lower() == ".csv":
        with path.open("r", newline="", encoding="utf-8") as fh:
            reader = csv.reader(fh)
            next(reader, None)  # skip header
            for row in reader:
                yield row
    else:
        workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
        sheet = workbook.active
        if sheet is None:
            raise ValueError("Workbook has no active sheet")
        first = True
        for row in sheet.iter_rows(values_only=True):
            if first:
                first = False
                continue
            yield ["" if cell is None else cell for cell in row]


def _next_output_path(root: Path, header: tuple[str, ...]) -> Path:
    signature = hashlib.sha1("||".join(header).encode("utf-8")).hexdigest()[:10]
    base_name = f"merged_{signature}.csv"
    candidate = root / base_name
    counter = 1
    while candidate.exists():
        candidate = root / f"merged_{signature}_{counter}.csv"
        counter += 1
    return candidate


@sdk.script
def main(ctx: sdk.ScriptContext, api: sdk.ScriptAPI) -> None:
    root = ctx.target_path
    if not root.exists() or not root.is_dir():
        raise ValueError(f"Target must be a directory: {root}")

    header_groups: dict[tuple[str, ...], list[Path]] = defaultdict(list)
    skipped: list[dict[str, str]] = []

    for entry in sorted(root.iterdir()):
        if not entry.is_file():
            continue
        if entry.suffix.lower() not in SUPPORTED_EXTS:
            continue
        try:
            header = tuple(_read_header(entry))
            if not header:
                raise ValueError("Missing header row")
            header_groups[header].append(entry)
        except Exception as exc:  # noqa: BLE001
            skipped.append({"file": str(entry), "reason": str(exc)})
            api.log("warn", f"Skipping {entry.name}: {exc}")

    merged_outputs: list[dict[str, object]] = []
    unmatched: list[str] = []

    for header, files in header_groups.items():
        if len(files) < 2:
            unmatched.extend(str(path) for path in files)
            continue

        output_path = _next_output_path(root, header)
        try:
            with output_path.open("w", newline="", encoding="utf-8") as fh:
                writer = csv.writer(fh)
                writer.writerow(header)
                for path in files:
                    try:
                        for row in _iter_rows(path):
                            writer.writerow(row)
                    except Exception as exc:  # noqa: BLE001
                        skipped.append({"file": str(path), "reason": str(exc)})
                        api.log("warn", f"Failed while reading {path.name}: {exc}")
            merged_outputs.append(
                {
                    "schema": list(header),
                    "output": str(output_path),
                    "sources": [str(p) for p in files],
                }
            )
            api.log(
                "info",
                f"Merged {len(files)} file(s) with schema {', '.join(header)} into {output_path.name}",
            )
        except Exception as exc:  # noqa: BLE001
            skipped.append({"file": str(output_path), "reason": f"write failed: {exc}"})
            api.log("error", f"Unable to write {output_path.name}: {exc}")
            if output_path.exists():
                output_path.unlink(missing_ok=True)

    api.emit_result(
        {
            "merged": merged_outputs,
            "skipped": skipped,
            "unmatched": unmatched,
        }
    )
    api.exit(code=0)


if __name__ == "__main__":
    main()
