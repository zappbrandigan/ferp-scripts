from __future__ import annotations

from collections.abc import Callable
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from ferp.fscp.scripts import sdk


def _norm(v: object) -> object:
    """Normalize values so trivial differences don't create noise."""
    if v is None:
        return ""
    if isinstance(v, str):
        return v.strip()
    return v


def _sheet_cells(
    ws_values, ws_source, stats: dict[str, int] | None = None
) -> dict[tuple[int, int], object]:
    """
    Return dict mapping (row, col) -> normalized value
    limited to ws.calculate_dimension() (used range).
    """
    cells = {}
    max_row_values = ws_values.max_row or 0
    max_col_values = ws_values.max_column or 0
    max_row_source = ws_source.max_row or 0
    max_col_source = ws_source.max_column or 0
    max_row = max(max_row_values, max_row_source)
    max_col = max(max_col_values, max_col_source)
    if max_row == 0 or max_col == 0:
        if stats is not None:
            stats.update(
                {
                    "max_row_values": max_row_values,
                    "max_col_values": max_col_values,
                    "max_row_source": max_row_source,
                    "max_col_source": max_col_source,
                    "cells_iterated": 0,
                    "values_nonempty": 0,
                    "source_nonempty": 0,
                    "source_fallback": 0,
                    "source_formulas": 0,
                    "final_nonempty": 0,
                }
            )
        return cells
    if stats is not None:
        stats.update(
            {
                "max_row_values": max_row_values,
                "max_col_values": max_col_values,
                "max_row_source": max_row_source,
                "max_col_source": max_col_source,
                "cells_iterated": 0,
                "values_nonempty": 0,
                "source_nonempty": 0,
                "source_fallback": 0,
                "source_formulas": 0,
                "final_nonempty": 0,
            }
        )
    for row_source in ws_source.iter_rows(
        min_row=1, max_row=max_row, min_col=1, max_col=max_col
    ):
        for cell_source in row_source:
            cell_values = ws_values.cell(row=cell_source.row, column=cell_source.column)
            value = cell_values.value
            used_source = value is None and cell_source.value is not None
            if used_source:
                value = cell_source.value
            if stats is not None:
                stats["cells_iterated"] += 1
                if cell_values.value is not None:
                    stats["values_nonempty"] += 1
                if cell_source.value is not None:
                    stats["source_nonempty"] += 1
                if used_source:
                    stats["source_fallback"] += 1
                if cell_source.data_type == "f":
                    stats["source_formulas"] += 1
                if value is not None and (not isinstance(value, str) or value.strip()):
                    stats["final_nonempty"] += 1
            cells[(cell_source.row, cell_source.column)] = _norm(value)
    return cells


def compare_excel(
    a_path: Path,
    b_path: Path,
    sheets: list[str] | None = None,
    log: Callable[[str], None] | None = None,
) -> tuple[list[tuple[str, int, int, object, object]], list[str], bool]:
    wb_a_values = load_workbook(a_path, data_only=True)
    wb_b_values = load_workbook(b_path, data_only=True)
    wb_a_source = load_workbook(a_path, data_only=False)
    wb_b_source = load_workbook(b_path, data_only=False)
    missing: list[str] = []
    compared_by_index = False
    try:
        pairs: list[tuple[str, str, str]] = []
        if log is not None:
            if not wb_a_values.sheetnames:
                log(f"Workbook '{a_path.name}' has no worksheets")
            if not wb_b_values.sheetnames:
                log(f"Workbook '{b_path.name}' has no worksheets")

        if sheets is None:
            common = sorted(set(wb_a_values.sheetnames) & set(wb_b_values.sheetnames))
            if common:
                pairs = [(name, name, name) for name in common]
            else:
                compared_by_index = True
                for a_name, b_name in zip(
                    wb_a_values.sheetnames, wb_b_values.sheetnames, strict=False
                ):
                    label = f"{a_name} vs {b_name}"
                    pairs.append((a_name, b_name, label))
        else:
            pairs = [(name, name, name) for name in sheets]
        if log is not None:
            log(
                "Workbook sheets | "
                f"a={wb_a_values.sheetnames} "
                f"b={wb_b_values.sheetnames} "
                f"pairs={[label for _, _, label in pairs]}"
            )
        diffs: list[tuple[str, int, int, object, object]] = []

        for a_name, b_name, label in pairs:
            if (
                a_name not in wb_a_values.sheetnames
                or b_name not in wb_b_values.sheetnames
            ):
                missing.append(label)
                continue

            ws_a_values = wb_a_values[a_name]
            ws_b_values = wb_b_values[b_name]
            ws_a_source = wb_a_source[a_name]
            ws_b_source = wb_b_source[b_name]

            a_stats: dict[str, int] = {}
            b_stats: dict[str, int] = {}
            a = _sheet_cells(ws_a_values, ws_a_source, stats=a_stats)
            b = _sheet_cells(ws_b_values, ws_b_source, stats=b_stats)
            if log is not None:
                log(f"Sheet stats | {label} | a={a_stats} b={b_stats}")

            coords = sorted(set(a.keys()) | set(b.keys()))
            if log is not None:
                log(f"Sheet '{label}' coordinate count: {len(coords)}")

            diffs_before = len(diffs)
            for r, c in coords:
                va = a.get((r, c), "")
                vb = b.get((r, c), "")
                if va != vb:
                    diffs.append((label, r, c, va, vb))
            if log is not None:
                log(f"Sheet '{label}' diffs: {len(diffs) - diffs_before}")
    finally:
        wb_a_values.close()
        wb_b_values.close()
        wb_a_source.close()
        wb_b_source.close()

    return diffs, missing, compared_by_index


def _write_xlsx(
    out_csv_path: Path, diffs: list[tuple[str, int, int, object, object]]
) -> None:
    out_csv_path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    if ws is None:
        ws = wb.create_sheet(title="diffs")
    else:
        ws.title = "diffs"

    headers = ["sheet", "col", "row", "a_value", "b_value"]
    ws.append(headers)

    header_fill = PatternFill(fill_type="solid", fgColor="D9D9D9")
    header_font = Font(bold=True)
    header_align = Alignment(vertical="center")
    for col_idx, _ in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align

    alt_fill = PatternFill(fill_type="solid", fgColor="F7F7F7")
    a_fill = PatternFill(fill_type="solid", fgColor="FCE4D6")
    b_fill = PatternFill(fill_type="solid", fgColor="E2F0D9")
    value_align = Alignment(vertical="top", wrap_text=True)

    for idx, (sheet, row, col, a_value, b_value) in enumerate(diffs, start=1):
        out_row = idx + 1
        ws.append([sheet, get_column_letter(col), row, a_value, b_value])
        if idx % 2 == 0:
            for col_idx in range(1, 6):
                ws.cell(row=out_row, column=col_idx).fill = alt_fill
        ws.cell(row=out_row, column=4).fill = a_fill
        ws.cell(row=out_row, column=5).fill = b_fill
        ws.cell(row=out_row, column=4).alignment = value_align
        ws.cell(row=out_row, column=5).alignment = value_align

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 8
    ws.column_dimensions["C"].width = 8
    ws.column_dimensions["D"].width = 40
    ws.column_dimensions["E"].width = 40

    wb.save(out_csv_path)


def _format_diffs(diffs: list[tuple[str, int, int, object, object]]) -> str:
    lines: list[str] = []
    for index, diff in enumerate(diffs):
        sheet, row, col, a_value, b_value = diff
        col_letter = get_column_letter(col)
        lines.append(f"{sheet}!{col_letter}{row}: {a_value} -> {b_value}")
    return "\n".join(lines).rstrip()


def _sample_cells(path: Path, max_cells: int = 5) -> list[tuple[str, str, object]]:
    wb = load_workbook(path, data_only=True)
    samples: list[tuple[str, str, object]] = []
    try:
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            max_row = ws.max_row or 0
            max_col = ws.max_column or 0
            if max_row == 0 or max_col == 0:
                continue
            for row in ws.iter_rows(
                min_row=1, max_row=max_row, min_col=1, max_col=max_col
            ):
                for cell in row:
                    if cell.column is None:
                        continue
                    value = cell.value
                    if value is None or (isinstance(value, str) and not value.strip()):
                        continue
                    samples.append(
                        (
                            sheet_name,
                            f"{get_column_letter(int(cell.column))}{cell.row}",
                            value,
                        )
                    )
                    if len(samples) >= max_cells:
                        return samples
    finally:
        wb.close()
    return samples


def _ensure_xlsx(path: Path, label: str) -> None:
    if path.suffix.lower() != ".xlsx":
        raise ValueError(f"{label} must be a .xlsx file: {path}")


def _resolve_xlsx(target_dir: Path, raw_value: str) -> Path:
    path = Path(raw_value)
    if not path.suffix:
        path = path.with_suffix(".xlsx")
    if not path.is_absolute():
        path = target_dir / path
    return path


@sdk.script
def main(ctx: sdk.ScriptContext, api: sdk.ScriptAPI) -> None:
    target_dir = ctx.target_path

    file_a = api.request_input(
        "First Excel file (relative to target directory)",
        id="excel_diff_file_a",
    )
    file_b = api.request_input(
        "Second Excel file (relative to target directory)",
        id="excel_diff_file_b",
    )
    sheets_raw = api.request_input(
        "Sheets to compare (comma-separated, optional)",
        id="excel_diff_sheets",
        default="",
    )

    if not file_a or not file_b:
        api.emit_result(
            {
                "_status": "error",
                "_title": "Error: Script Canceled",
                "Message": "User failed to enter two Excel file paths.",
            }
        )
        api.exit(code=1)
        return

    a_path = _resolve_xlsx(target_dir, file_a)
    b_path = _resolve_xlsx(target_dir, file_b)

    if not a_path.exists():
        api.emit_result(
            {
                "_status": "error",
                "_title": "Error: File Not Found",
                "Provided File": str(a_path),
            }
        )
        return
    if not b_path.exists():
        api.emit_result(
            {
                "_status": "error",
                "_title": "Error: File Not Found",
                "Provided File": str(b_path),
            }
        )
        return

    _ensure_xlsx(a_path, "file_a")
    _ensure_xlsx(b_path, "file_b")

    sheets = [s.strip() for s in sheets_raw.split(",") if s.strip()] or None

    for label, path in (("file_a", a_path), ("file_b", b_path)):
        samples = _sample_cells(path)
        if samples:
            api.log("info", f"{label} sample cells: {samples}")
        else:
            api.log("warn", f"{label} has no non-empty cells detected")

    api.log(
        "info",
        f"Comparing '{a_path.name}' vs '{b_path.name}' | sheets={sheets or 'auto'}",
    )

    diffs, missing, compared_by_index = compare_excel(
        a_path, b_path, sheets=sheets, log=lambda msg: api.log("debug", msg)
    )

    xlsx_path: Path | None = None
    output_dir = a_path.parent
    if a_path.parent != b_path.parent:
        api.log(
            "warn",
            (
                "Input files are in different directories; writing XLSX next to "
                f"the first file: {a_path.parent}"
            ),
        )
    xlsx_path = output_dir / f"{a_path.stem}_vs_{b_path.stem}_diff.xlsx"
    _write_xlsx(xlsx_path, diffs)

    if missing:
        api.log("warn", f"Skipped missing sheets: {', '.join(missing)}")
    if compared_by_index:
        api.log(
            "warn",
            (
                "No matching sheet names found; compared sheets by index order instead. "
                "Provide explicit sheet names to override."
            ),
        )

    api.emit_result(
        {
            "_title": "Excel Diff Results",
            "Diffs Found": len(diffs),
            "Missing Sheets": missing if missing else "None",
            "Compared by Index": compared_by_index,
            "File Path": str(xlsx_path) if xlsx_path else None,
            "Results": f"\n{_format_diffs(diffs)}",
        }
    )


if __name__ == "__main__":
    main()
