from __future__ import annotations

import re
from collections import Counter
from dataclasses import dataclass, field
from pathlib import Path
from typing import Iterable, Sequence, TypeAlias, TypedDict

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.worksheet import Worksheet

from ferp.fscp.scripts import sdk
from ferp.fscp.scripts.common import FerpXmpMetadata, build_destination, read_pdf_ferp_metadata


@dataclass(frozen=True)
class CueSheetRow:
    relative_path: str
    date_range: str
    deal: str
    catalog: str
    pd_code: str
    film_or_series: str
    revision_status: str
    cue_sheet: str
    production_title: str
    episode_title: str
    season_number: str
    episode_number: str
    territory_code: str
    territory: str


@dataclass(frozen=True)
class CueSheetRecord:
    path_row: CueSheetRow
    administrator: str
    xmp_catalog_code: str
    data_added_date: str
    stamp_spec_version: str
    document_id: str
    instance_id: str
    publishers: tuple[str, ...]
    effective_dates: tuple[str, ...]
    xmp_territories: tuple[str, ...]
    agreement_count: int
    effective_date_count: int
    xmp_present: bool
    ferp_metadata_present: bool
    metadata_error: str


class ReportModeResponse(TypedDict):
    value: str
    report_mode: str


SummaryRow: TypeAlias = tuple[str, object]
CueSheetDetailRow: TypeAlias = tuple[
    str,
    str,
    str,
    str,
    str,
    str,
    str,
    str,
    str,
    str,
    str,
    str,
    str,
    str,
    str,
    str,
    str,
    int,
    int,
    str,
    str,
    str,
    str,
]
CatalogInsightRow: TypeAlias = tuple[
    str,
    str,
    int,
    int,
    str,
    int,
    int,
    int,
    int,
    int,
]
PublisherInsightRow: TypeAlias = tuple[
    str,
    int,
    int,
    str,
    int,
]
QualityRow: TypeAlias = tuple[str, str, str, str, str, str, str]


@dataclass
class CatalogInsightBucket:
    cue_sheets: int = 0
    revisions: int = 0
    new_files: int = 0
    films: int = 0
    series: int = 0
    productions: set[str] = field(default_factory=set)
    publishers: set[str] = field(default_factory=set)


@dataclass
class PublisherInsightBucket:
    cue_sheets: int = 0
    revisions: int = 0
    productions: set[str] = field(default_factory=set)


_PRODUCTION_DELIM = "   "
_EPISODE_DELIM = "  "
_DATEISH_EPISODE_NUMBER_RE = re.compile(
    r"^\d{6,8}(?:-\d+[A-Za-z]?)?$|^\d{1,2}[-/]\d{1,2}[-/]\d{2,4}$"
)
_EPISODE_CODE_RE = re.compile(r"^\d{4,5}$")
_ISO_DATE_RE = re.compile(r"^\d{4}-\d{2}-\d{2}$")
_MONTHLY_REPORT_MODE = "Monthly Report"
_WEEKLY_REPORT_MODE = "Weekly Report"
_STATISTICAL_REPORT_MODE = "Statistical Report"
_MONTHLY_REPORT_SUFFIX = "monthly_report"
_WEEKLY_REPORT_SUFFIX = "weekly_report"
_STATISTICAL_REPORT_SUFFIX = "statistical_report"


def _normalize_text(value: str) -> str:
    return " ".join(value.split()).strip()


def _sorted_unique(values: Iterable[str]) -> tuple[str, ...]:
    deduped: list[str] = []
    seen: set[str] = set()
    for raw in values:
        value = _normalize_text(raw)
        if not value or value in seen:
            continue
        seen.add(value)
        deduped.append(value)
    return tuple(sorted(deduped, key=str.lower))


def _sorted_dates(values: Iterable[str]) -> tuple[str, ...]:
    deduped = _sorted_unique(values)
    valid = [value for value in deduped if _ISO_DATE_RE.fullmatch(value)]
    invalid = [value for value in deduped if not _ISO_DATE_RE.fullmatch(value)]
    return tuple(sorted(valid) + invalid)


def _date_bounds(values: Iterable[str]) -> tuple[str, str]:
    valid_dates = [value for value in _sorted_dates(values) if _ISO_DATE_RE.fullmatch(value)]
    if not valid_dates:
        return "", ""
    return valid_dates[0], valid_dates[-1]


def _display_join(values: Iterable[str]) -> str:
    return ", ".join(values)


def _yes_no(value: bool) -> str:
    return "Yes" if value else "No"


def _ratio(numerator: int, denominator: int) -> str:
    if denominator <= 0:
        return ""
    return f"{(numerator / denominator):.1%}"


def _record_sort_key(record: CueSheetRecord) -> tuple[str, ...]:
    row = record.path_row
    return (
        row.date_range.lower(),
        row.deal.lower(),
        row.catalog.lower(),
        row.film_or_series.lower(),
        row.revision_status.lower(),
        row.production_title.lower(),
        row.episode_title.lower(),
        row.season_number.lower(),
        row.episode_number.lower(),
        row.territory_code.lower(),
        row.relative_path.lower(),
    )


def _path_row_sort_key(row: CueSheetRow) -> tuple[str, ...]:
    return (
        row.date_range.lower(),
        row.deal.lower(),
        row.catalog.lower(),
        row.pd_code.lower(),
        row.film_or_series.lower(),
        row.revision_status.lower(),
        row.production_title.lower(),
        row.episode_title.lower(),
        row.season_number.lower(),
        row.episode_number.lower(),
        row.territory_code.lower(),
        row.territory.lower(),
        row.relative_path.lower(),
    )


def _catalog_match(path_catalog: str, xmp_catalog_code: str) -> str:
    if not path_catalog or not xmp_catalog_code:
        return ""
    return _yes_no(path_catalog.casefold() == xmp_catalog_code.casefold())


def _parse_series_title_fields(stem: str) -> tuple[str, str, str, str]:
    if _PRODUCTION_DELIM not in stem:
        return stem, "", "", ""

    production_title, remainder = stem.split(_PRODUCTION_DELIM, 1)
    production_title = production_title.strip()
    remainder = remainder.strip()
    if not remainder or "Ep No." not in remainder:
        return production_title, "", "", ""

    episode_title = ""
    episode_info = remainder
    if _EPISODE_DELIM in remainder:
        possible_episode_title, possible_episode_info = remainder.rsplit(
            _EPISODE_DELIM, 1
        )
        if "Ep No." in possible_episode_info:
            episode_title = possible_episode_title.strip()
            episode_info = possible_episode_info.strip()

    episode_info = episode_info.strip()
    if "Ep No." not in episode_info:
        return production_title, episode_title, "", ""

    _, episode_number = episode_info.split("Ep No.", 1)
    episode_number = episode_number.strip()
    if not episode_number or _DATEISH_EPISODE_NUMBER_RE.fullmatch(episode_number):
        return production_title, episode_title, "", ""

    season_number = ""
    if _EPISODE_CODE_RE.fullmatch(episode_number):
        if len(episode_number) == 4:
            season_number = episode_number[0]
        else:
            season_number = episode_number[:2]
        episode_number = str(int(episode_number[-3:]))

    return production_title, episode_title, season_number, episode_number


def _parse_title_fields(stem: str, media_type: str) -> tuple[str, str, str, str]:
    if media_type == "film":
        return stem, "", "", ""
    return _parse_series_title_fields(stem)


def _parse_cue_sheet_path(pdf_path: Path, root: Path) -> CueSheetRow | None:
    relative_path = pdf_path.relative_to(root)
    relative_parts = relative_path.parts
    if len(relative_parts) < 4:
        return None

    filename = relative_parts[-1]
    catalog_folder = relative_parts[-2]
    media_type = relative_parts[-3].lower()
    territory_code = relative_parts[-4]

    if media_type not in {"film", "tv"}:
        return None

    if len(relative_parts) >= 5 and relative_parts[-5] == "_REV":
        date_range = root.name if len(relative_parts) == 5 else relative_parts[-6]
        revision_status = "Revision"
    elif len(relative_parts) == 4:
        date_range = root.name
        revision_status = "New"
    else:
        date_range = relative_parts[-5]
        revision_status = "New"

    if not date_range or date_range == "_REV":
        return None
    if not catalog_folder:
        return None
    if not territory_code:
        return None
    if not filename.lower().endswith(".pdf"):
        return None

    if " - " in catalog_folder:
        catalog, deal = catalog_folder.split(" - ", 1)
    else:
        catalog = catalog_folder
        deal = ""

    film_or_series = "Film" if media_type == "film" else "Series"
    title_stem = Path(filename).stem
    production_title, episode_title, season_number, episode_number = (
        _parse_title_fields(title_stem, media_type)
    )

    return CueSheetRow(
        relative_path=str(relative_path),
        date_range=date_range,
        deal=deal,
        catalog=catalog,
        pd_code="",
        film_or_series=film_or_series,
        revision_status=revision_status,
        cue_sheet=title_stem,
        production_title=production_title,
        episode_title=episode_title,
        season_number=season_number,
        episode_number=episode_number,
        territory_code=territory_code,
        territory="",
    )


def _collect_pdfs(root: Path, api: sdk.ScriptAPI) -> list[Path]:
    files: list[Path] = []
    for path in root.rglob("*.pdf"):
        api.check_cancel()
        if path.is_file() and not path.name.startswith((".", "~$")):
            files.append(path)
    return sorted(files)


def _record_from_metadata(
    path_row: CueSheetRow,
    metadata: FerpXmpMetadata | None,
    metadata_error: str = "",
) -> CueSheetRecord:
    publishers: list[str] = []
    effective_dates: list[str] = []
    xmp_territories: list[str] = []
    agreement_count = 0
    effective_date_count = 0
    if metadata is not None:
        agreement_count = len(metadata.agreements)
        for agreement in metadata.agreements:
            publishers.extend(agreement.publishers)
            effective_date_count += len(agreement.effective_dates)
            for effective_date in agreement.effective_dates:
                if effective_date.date:
                    effective_dates.append(effective_date.date)
                xmp_territories.extend(effective_date.territories)

    ferp_metadata_present = bool(
        metadata
        and (
            metadata.administrator
            or metadata.catalog_code
            or metadata.data_added_date
            or metadata.stamp_spec_version
            or metadata.agreements
        )
    )

    return CueSheetRecord(
        path_row=path_row,
        administrator=metadata.administrator if metadata else "",
        xmp_catalog_code=metadata.catalog_code if metadata else "",
        data_added_date=metadata.data_added_date if metadata else "",
        stamp_spec_version=metadata.stamp_spec_version if metadata else "",
        document_id=(metadata.document_id or "") if metadata else "",
        instance_id=(metadata.instance_id or "") if metadata else "",
        publishers=_sorted_unique(publishers),
        effective_dates=_sorted_dates(effective_dates),
        xmp_territories=_sorted_unique(xmp_territories),
        agreement_count=agreement_count,
        effective_date_count=effective_date_count,
        xmp_present=metadata is not None,
        ferp_metadata_present=ferp_metadata_present,
        metadata_error=_normalize_text(metadata_error),
    )


def _build_table_sheet(
    wb: Workbook,
    title: str,
    headers: list[str],
    rows: Sequence[Sequence[object]],
    *,
    table_name: str,
) -> None:
    ws = wb.create_sheet(title=title)

    header_font = Font(bold=True)
    header_alignment = Alignment(horizontal="center", vertical="center")
    data_alignment = Alignment(vertical="center")

    ws.append(headers)
    for column_index, _header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=column_index)
        cell.font = header_font
        cell.alignment = header_alignment

    for row in rows:
        ws.append(list(row))

    ws.freeze_panes = "A2"

    if rows:
        last_row = len(rows) + 1
        last_col = len(headers)
        table_ref = f"A1:{get_column_letter(last_col)}{last_row}"
        table = Table(displayName=table_name, ref=table_ref)
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleLight1",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        ws.add_table(table)

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = data_alignment

    for column_index, header in enumerate(headers, start=1):
        max_length = len(str(header))
        for row in rows:
            value = row[column_index - 1] if column_index - 1 < len(row) else ""
            max_length = max(max_length, len(str(value)))
        ws.column_dimensions[get_column_letter(column_index)].width = min(
            max(max_length + 2, 12), 60
        )


def _require_active_sheet(workbook: Workbook) -> Worksheet:
    worksheet = workbook.active
    if not isinstance(worksheet, Worksheet):
        worksheet = workbook.create_sheet()
    return worksheet


def _write_report(xlsx_path: Path, rows: list[CueSheetRow]) -> None:
    xlsx_path.parent.mkdir(parents=True, exist_ok=True)

    headers = [
        "Deal",
        "Catalog",
        "PD Code",
        "Film or Series",
        "Revision Status",
        "Production Title",
        "Episode Title",
        "Season Number",
        "Episode Number",
        "Territory Code",
        "Territory",
    ]

    workbook = Workbook()
    default_sheet = _require_active_sheet(workbook)
    default_sheet.title = "Cue Sheet Summary"

    header_font = Font(bold=True)
    header_alignment = Alignment(horizontal="center", vertical="center")
    data_alignment = Alignment(vertical="center")

    default_sheet.append(headers)
    for column_index, _header in enumerate(headers, start=1):
        cell = default_sheet.cell(row=1, column=column_index)
        cell.font = header_font
        cell.alignment = header_alignment

    for row in rows:
        default_sheet.append(
            [
                row.deal,
                row.catalog,
                row.pd_code,
                row.film_or_series,
                row.revision_status,
                row.production_title,
                row.episode_title,
                row.season_number,
                row.episode_number,
                row.territory_code,
                row.territory,
            ]
        )

    if rows:
        last_row = len(rows) + 1
        last_col = len(headers)
        table_ref = f"A1:{get_column_letter(last_col)}{last_row}"
        table = Table(displayName="CueSheetSummary", ref=table_ref)
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleLight1",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        default_sheet.add_table(table)

    column_widths = {
        "A": 40,
        "B": 18,
        "C": 18,
        "D": 16,
        "E": 18,
        "F": 40,
        "G": 32,
        "H": 16,
        "I": 20,
        "J": 18,
        "K": 18,
    }
    for column, width in column_widths.items():
        default_sheet.column_dimensions[column].width = width

    for row in default_sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = data_alignment

    workbook.save(xlsx_path)


def _write_monthly_report(xlsx_path: Path, rows: list[CueSheetRow]) -> None:
    _write_report(xlsx_path, rows)


def _write_weekly_report(xlsx_path: Path, rows: list[CueSheetRow]) -> None:
    xlsx_path.parent.mkdir(parents=True, exist_ok=True)

    headers = [
        "Deal",
        "Catalog",
        "Film or Series",
        "Revision Status",
        "Cue Sheet",
        "Territory Code",
        "Territory",
    ]

    workbook = Workbook()
    default_sheet = _require_active_sheet(workbook)
    default_sheet.title = "Cue Sheet Summary"

    header_font = Font(bold=True)
    header_alignment = Alignment(horizontal="center", vertical="center")
    data_alignment = Alignment(vertical="center")

    default_sheet.append(headers)
    for column_index, _header in enumerate(headers, start=1):
        cell = default_sheet.cell(row=1, column=column_index)
        cell.font = header_font
        cell.alignment = header_alignment

    for row in rows:
        default_sheet.append(
            [
                row.deal,
                row.catalog,
                row.film_or_series,
                row.revision_status,
                row.cue_sheet,
                row.territory_code,
                row.territory,
            ]
        )

    if rows:
        last_row = len(rows) + 1
        last_col = len(headers)
        table_ref = f"A1:{get_column_letter(last_col)}{last_row}"
        table = Table(displayName="CueSheetSummary", ref=table_ref)
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleLight1",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        default_sheet.add_table(table)

    column_widths = {
        "A": 40,
        "B": 18,
        "C": 16,
        "D": 18,
        "E": 56,
        "F": 18,
        "G": 18,
    }
    for column, width in column_widths.items():
        default_sheet.column_dimensions[column].width = width

    for row in default_sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = data_alignment

    workbook.save(xlsx_path)


def _build_summary_rows(
    records: list[CueSheetRecord],
    skipped: list[str],
) -> list[SummaryRow]:
    all_publishers = sorted({publisher for record in records for publisher in record.publishers})
    all_stamp_dates = [record.data_added_date for record in records if record.data_added_date]
    all_effective_dates = [
        effective_date
        for record in records
        for effective_date in record.effective_dates
        if effective_date
    ]
    earliest_stamp_date, latest_stamp_date = _date_bounds(all_stamp_dates)
    earliest_effective_date, latest_effective_date = _date_bounds(all_effective_dates)
    document_ids = [record.document_id for record in records if record.document_id]
    duplicate_document_ids = sum(1 for count in Counter(document_ids).values() if count > 1)

    return [
        ("Total PDFs Scanned", len(records) + len(skipped)),
        ("Recognized Cue Sheets", len(records)),
        ("Skipped PDFs", len(skipped)),
        ("Cue Sheets With FERP Metadata", sum(1 for record in records if record.ferp_metadata_present)),
        ("Cue Sheets With XMP DocumentID", sum(1 for record in records if record.document_id)),
        ("Cue Sheets Missing FERP Metadata", sum(1 for record in records if not record.ferp_metadata_present)),
        ("Cue Sheets Missing XMP Catalog Code", sum(1 for record in records if not record.xmp_catalog_code)),
        ("Cue Sheets Missing Administrator", sum(1 for record in records if not record.administrator)),
        ("Cue Sheets Missing Stamp Date", sum(1 for record in records if not record.data_added_date)),
        ("Cue Sheets Missing Agreements", sum(1 for record in records if record.agreement_count == 0)),
        ("Cue Sheets Missing Publishers", sum(1 for record in records if not record.publishers)),
        ("Cue Sheets Missing Effective Dates", sum(1 for record in records if not record.effective_dates)),
        ("Cue Sheets Missing Territories", sum(1 for record in records if not record.xmp_territories)),
        ("Cue Sheets With Metadata Read Errors", sum(1 for record in records if record.metadata_error)),
        ("Revision PDFs", sum(1 for record in records if record.path_row.revision_status == "Revision")),
        ("New PDFs", sum(1 for record in records if record.path_row.revision_status == "New")),
        ("Film PDFs", sum(1 for record in records if record.path_row.film_or_series == "Film")),
        ("Series PDFs", sum(1 for record in records if record.path_row.film_or_series == "Series")),
        ("Unique Path Catalogs", len({record.path_row.catalog for record in records if record.path_row.catalog})),
        ("Unique XMP Catalog Codes", len({record.xmp_catalog_code for record in records if record.xmp_catalog_code})),
        ("Unique Publishers", len(all_publishers)),
        ("Duplicate DocumentIDs", duplicate_document_ids),
        ("Earliest Stamp Date", earliest_stamp_date),
        ("Latest Stamp Date", latest_stamp_date),
        ("Earliest Effective Date", earliest_effective_date),
        ("Latest Effective Date", latest_effective_date),
    ]


def _build_cue_sheet_rows(records: list[CueSheetRecord]) -> list[CueSheetDetailRow]:
    rows: list[CueSheetDetailRow] = []
    for record in sorted(records, key=_record_sort_key):
        row = record.path_row
        rows.append(
            (
                row.date_range,
                row.deal,
                row.catalog,
                record.xmp_catalog_code,
                _catalog_match(row.catalog, record.xmp_catalog_code),
                row.film_or_series,
                row.revision_status,
                row.territory_code,
                row.cue_sheet,
                row.production_title,
                row.episode_title,
                row.season_number,
                row.episode_number,
                record.data_added_date,
                record.stamp_spec_version,
                record.document_id,
                record.instance_id,
                record.agreement_count,
                len(record.publishers),
                _display_join(record.publishers),
                _display_join(record.xmp_territories),
                _yes_no(record.ferp_metadata_present),
                record.metadata_error,
            )
        )
    return rows


def _build_catalog_insight_rows(records: list[CueSheetRecord]) -> list[CatalogInsightRow]:
    aggregates: dict[tuple[str, str], CatalogInsightBucket] = {}
    for record in records:
        key = (record.path_row.catalog, record.xmp_catalog_code)
        bucket = aggregates.setdefault(key, CatalogInsightBucket())
        bucket.cue_sheets += 1
        if record.path_row.revision_status == "Revision":
            bucket.revisions += 1
        else:
            bucket.new_files += 1
        if record.path_row.film_or_series == "Film":
            bucket.films += 1
        else:
            bucket.series += 1
        if record.path_row.production_title:
            bucket.productions.add(record.path_row.production_title)
        bucket.publishers.update(record.publishers)

    rows: list[CatalogInsightRow] = []
    for (path_catalog, xmp_catalog), bucket in aggregates.items():
        cue_sheets = bucket.cue_sheets
        revisions = bucket.revisions
        rows.append(
            (
                path_catalog,
                xmp_catalog,
                cue_sheets,
                revisions,
                _ratio(revisions, cue_sheets),
                bucket.new_files,
                bucket.films,
                bucket.series,
                len(bucket.productions),
                len(bucket.publishers),
            )
        )

    rows.sort(key=lambda item: (-item[2], item[0].lower(), item[1].lower()))
    return rows


def _build_publisher_insight_rows(records: list[CueSheetRecord]) -> list[PublisherInsightRow]:
    aggregates: dict[str, PublisherInsightBucket] = {}
    for record in records:
        for publisher in record.publishers:
            bucket = aggregates.setdefault(publisher, PublisherInsightBucket())
            bucket.cue_sheets += 1
            if record.path_row.revision_status == "Revision":
                bucket.revisions += 1
            if record.path_row.production_title:
                bucket.productions.add(record.path_row.production_title)

    rows: list[PublisherInsightRow] = []
    for publisher, bucket in aggregates.items():
        cue_sheets = bucket.cue_sheets
        revisions = bucket.revisions
        rows.append(
            (
                publisher,
                cue_sheets,
                revisions,
                _ratio(revisions, cue_sheets),
                len(bucket.productions),
            )
        )

    rows.sort(key=lambda item: (-item[1], item[0].lower()))
    return rows


def _build_quality_rows(
    records: list[CueSheetRecord],
    skipped: list[str],
) -> list[QualityRow]:
    duplicate_ids = {
        value
        for value, count in Counter(
            record.document_id for record in records if record.document_id
        ).items()
        if count > 1
    }
    rows: list[QualityRow] = []

    for relative_path in skipped:
        rows.append((relative_path, "Unexpected Folder Layout", "", "", "", "", ""))

    for record in records:
        row = record.path_row

        def add_issue(issue: str, detail: str = "") -> None:
            rows.append(
                (
                    row.relative_path,
                    issue,
                    detail,
                    row.catalog,
                    record.xmp_catalog_code,
                    row.revision_status,
                    record.document_id,
                )
            )

        if record.metadata_error:
            add_issue("Metadata Read Failed", record.metadata_error)
        if not record.ferp_metadata_present:
            add_issue("Missing FERP Metadata")
        if not record.document_id:
            add_issue("Missing XMP DocumentID")
        if not record.administrator:
            add_issue("Missing Administrator")
        if not record.xmp_catalog_code:
            add_issue("Missing XMP Catalog Code")
        elif row.catalog and row.catalog.casefold() != record.xmp_catalog_code.casefold():
            add_issue("Catalog Mismatch", f"path={row.catalog}; xmp={record.xmp_catalog_code}")
        if not record.data_added_date:
            add_issue("Missing Stamp Date")
        elif not _ISO_DATE_RE.fullmatch(record.data_added_date):
            add_issue("Invalid Stamp Date", record.data_added_date)
        if record.agreement_count == 0:
            add_issue("Missing Agreements")
        if not record.publishers:
            add_issue("Missing Publishers")
        if not record.effective_dates:
            add_issue("Missing Effective Dates")
        else:
            invalid_dates = [value for value in record.effective_dates if not _ISO_DATE_RE.fullmatch(value)]
            if invalid_dates:
                add_issue("Invalid Effective Dates", _display_join(invalid_dates))
        if not record.xmp_territories:
            add_issue("Missing Territories")
        if record.document_id and record.document_id in duplicate_ids:
            add_issue("Duplicate DocumentID", record.document_id)

    rows.sort(key=lambda item: (str(item[1]).lower(), str(item[0]).lower()))
    return rows


def _write_statistical_report(
    xlsx_path: Path,
    records: list[CueSheetRecord],
    skipped: list[str],
) -> None:
    xlsx_path.parent.mkdir(parents=True, exist_ok=True)

    workbook = Workbook()
    default_sheet = workbook.active
    if default_sheet is not None:
        workbook.remove(default_sheet)

    catalog_rows = _build_catalog_insight_rows(records)
    publisher_rows = _build_publisher_insight_rows(records)
    _build_table_sheet(
        workbook,
        "Summary",
        ["Metric", "Value"],
        _build_summary_rows(records, skipped),
        table_name="SummaryMetrics",
    )
    _build_table_sheet(
        workbook,
        "Cue Sheets",
        [
            "Folder Date Range",
            "Deal",
            "Path Catalog",
            "XMP Catalog Code",
            "Catalog Match",
            "Film or Series",
            "Revision Status",
            "Territory Code",
            "Cue Sheet",
            "Production Title",
            "Episode Title",
            "Season Number",
            "Episode Number",
            "Stamp Date",
            "Stamp Spec Version",
            "DocumentID",
            "InstanceID",
            "Agreement Count",
            "Publisher Count",
            "Publishers",
            "XMP Territories",
            "Has FERP Metadata",
            "Metadata Error",
        ],
        _build_cue_sheet_rows(records),
        table_name="CueSheetDetails",
    )
    _build_table_sheet(
        workbook,
        "Catalog Insights",
        [
            "Path Catalog",
            "XMP Catalog Code",
            "Cue Sheets",
            "Revisions",
            "Revision Rate",
            "New Files",
            "Films",
            "Series",
            "Productions",
            "Unique Publishers",
        ],
        catalog_rows,
        table_name="CatalogInsights",
    )
    _build_table_sheet(
        workbook,
        "Publisher Insights",
        [
            "Publisher",
            "Cue Sheets",
            "Revisions",
            "Revision Rate",
            "Productions",
        ],
        publisher_rows,
        table_name="PublisherInsights",
    )
    _build_table_sheet(
        workbook,
        "Quality Checks",
        [
            "Relative Path",
            "Issue",
            "Detail",
            "Path Catalog",
            "XMP Catalog Code",
            "Revision Status",
            "DocumentID",
        ],
        _build_quality_rows(records, skipped),
        table_name="QualityChecks",
    )

    workbook.save(xlsx_path)


def _report_output_stem(target: Path, report_mode: str) -> str:
    suffix_by_mode = {
        _MONTHLY_REPORT_MODE: _MONTHLY_REPORT_SUFFIX,
        _WEEKLY_REPORT_MODE: _WEEKLY_REPORT_SUFFIX,
        _STATISTICAL_REPORT_MODE: _STATISTICAL_REPORT_SUFFIX,
    }
    suffix = suffix_by_mode.get(report_mode, _WEEKLY_REPORT_SUFFIX)
    return f"{target.name}_cue_sheet_{suffix}"


@sdk.script
def main(ctx: sdk.ScriptContext, api: sdk.ScriptAPI) -> None:
    target = ctx.target_path
    if not target.is_dir():
        api.emit_result(
            {
                "_status": "warn",
                "_title": "Cue Sheet Report Canceled",
                "Info": "Select a directory to build the cue sheet summary.",
            }
        )
        return

    payload = api.request_input_json(
        "Select cue sheet report type",
        id="ferp_report_cue_sheet_directory_mode",
        fields=[
            {
                "id": "report_mode",
                "type": "select",
                "label": "Report Type",
                "options": [
                    _WEEKLY_REPORT_MODE,
                    _MONTHLY_REPORT_MODE,
                    _STATISTICAL_REPORT_MODE,
                ],
                "default": _WEEKLY_REPORT_MODE,
            }
        ],
        show_text_input=False,
        payload_type=ReportModeResponse,
    )
    report_mode = str(payload.get("report_mode") or _MONTHLY_REPORT_MODE).strip()

    pdf_files = _collect_pdfs(target, api)
    if not pdf_files:
        api.emit_result(
            {
                "_status": "warn",
                "_title": "Cue Sheet Report Complete",
                "Report Type": report_mode,
                "Rows Written": 0,
                "Skipped PDFs": 0,
                "Info": "No PDF files were found under the selected directory.",
            }
        )
        return

    rows: list[CueSheetRow] = []
    records: list[CueSheetRecord] = []
    skipped: list[str] = []
    total_files = len(pdf_files)

    for index, pdf_path in enumerate(pdf_files, start=1):
        api.check_cancel()
        parsed = _parse_cue_sheet_path(pdf_path, target)
        if parsed is None:
            skipped.append(str(pdf_path.relative_to(target)))
        else:
            rows.append(parsed)
            if report_mode == _STATISTICAL_REPORT_MODE:
                try:
                    metadata = read_pdf_ferp_metadata(pdf_path)
                    metadata_error = ""
                except Exception as exc:  # noqa: BLE001
                    metadata = None
                    metadata_error = str(exc)
                    api.log(
                        "warn",
                        f"Failed to read XMP metadata for '{pdf_path.name}': {exc}",
                    )
                records.append(_record_from_metadata(parsed, metadata, metadata_error))
        api.progress(current=index, total=total_files, unit="files")

    rows.sort(key=_path_row_sort_key)
    records.sort(key=_record_sort_key)

    out_path = build_destination(
        target.parent,
        _report_output_stem(target, report_mode),
        ".xlsx",
    )
    if report_mode == _MONTHLY_REPORT_MODE:
        _write_monthly_report(out_path, rows)
    elif report_mode == _WEEKLY_REPORT_MODE:
        _write_weekly_report(out_path, rows)
    else:
        _write_statistical_report(out_path, records, skipped)

    if skipped:
        preview = ", ".join(skipped[:5])
        api.log(
            "warn",
            f"Skipped {len(skipped)} PDF(s) that did not match the expected folder layout: {preview}",
        )

    result: dict[str, object] = {
        "_title": "Cue Sheet Report Complete",
        "Report Type": report_mode,
        "XLSX Path": str(out_path),
        "Rows Written": len(records) if report_mode == _STATISTICAL_REPORT_MODE else len(rows),
        "Skipped PDFs": len(skipped),
    }
    if report_mode == _STATISTICAL_REPORT_MODE:
        result["Metadata Read Errors"] = sum(1 for record in records if record.metadata_error)
        result["Cue Sheets With FERP Metadata"] = sum(
            1 for record in records if record.ferp_metadata_present
        )

    api.emit_result(result)


if __name__ == "__main__":
    main()
