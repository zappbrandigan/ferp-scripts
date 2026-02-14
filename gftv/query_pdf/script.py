from __future__ import annotations

import re
from collections import defaultdict
from dataclasses import dataclass
from pathlib import Path
from typing import Callable, TypedDict

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from pypdf import PdfReader

from ferp.fscp.scripts import sdk
from ferp.fscp.scripts.common import collect_files


class QueryOptions(TypedDict):
    value: str
    recursive: bool
    case_sensitive: bool


@dataclass(frozen=True)
class QuerySpec:
    label: str
    pattern: re.Pattern[str]


def _compile_pattern(
    query: str,
    use_regex: bool,
    case_sensitive: bool,
    api: sdk.ScriptAPI,
) -> re.Pattern[str] | None:
    flags = 0 if case_sensitive else re.IGNORECASE
    try:
        if use_regex:
            return re.compile(query, flags)
        return re.compile(re.escape(query), flags)
    except re.error as exc:
        api.log("error", f"Invalid regular expression '{query}': {exc}")
        return None


def _parse_queries(
    raw_query: str,
    *,
    case_sensitive: bool,
    api: sdk.ScriptAPI,
) -> tuple[list[QuerySpec], list[str]]:
    raw_lines = [line.strip() for line in raw_query.splitlines()]
    lines = [line for line in raw_lines if line]
    seen: set[tuple[str, bool]] = set()
    queries: list[QuerySpec] = []
    invalid: list[str] = []

    for line in lines:
        is_wrapped_regex = line.startswith("/") and line.endswith("/") and len(line) > 1
        use_regex = is_wrapped_regex
        query_text = line[1:-1] if is_wrapped_regex else line
        key = (query_text, use_regex)
        if key in seen:
            continue
        seen.add(key)
        pattern = _compile_pattern(query_text, use_regex, case_sensitive, api)
        if pattern is None:
            invalid.append(line)
            continue
        queries.append(QuerySpec(label=line, pattern=pattern))

    return queries, invalid


def _process_pdf(
    pdf_path: Path,
    root: Path,
    queries: list[QuerySpec],
    context_chars: int,
    api: sdk.ScriptAPI,
    check_cancel: Callable[[], None] | None = None,
) -> list[dict[str, object]]:
    reader = PdfReader(str(pdf_path))
    if reader.is_encrypted:
        api.log("warn", f"Skipping encrypted PDF: {pdf_path.name}")
        return []

    pdf_matches: list[dict[str, object]] = []
    relative_path = pdf_path.relative_to(root)

    for page_number, page in enumerate(reader.pages, start=1):
        if check_cancel is not None:
            check_cancel()
        try:
            text = page.extract_text() or ""
        except Exception as exc:  # noqa: BLE001
            api.log(
                "warn",
                f"Text extraction failed for '{pdf_path.name}' (page {page_number}): {exc}",
            )
            continue

        if not text:
            continue

        for query in queries:
            for match in query.pattern.finditer(text):
                if check_cancel is not None:
                    check_cancel()
                span = match.span()
                context = _extract_context(text, span[0], span[1], context_chars)
                pdf_matches.append(
                    {
                        "file_name": pdf_path.name,
                        "relative_path": str(relative_path),
                        "query": query.label,
                        "page_number": page_number,
                        "match_text": match.group(0),
                        "context_excerpt": context,
                    }
                )

    if not pdf_matches:
        api.log("info", f"No matches found in {pdf_path.name}")

    return pdf_matches


def _extract_context(text: str, start: int, end: int, context_chars: int) -> str:
    left = max(0, start - context_chars)
    right = min(len(text), end + context_chars)
    snippet = text[left:right].replace("\n", " ").strip()
    return snippet


def _write_xlsx(xlsx_path: Path, rows: list[dict[str, object]]) -> None:
    headers = [
        "file_name",
        "relative_path",
        "query",
        "page_number",
        "match_text",
        "context_excerpt",
    ]
    wb = Workbook()
    ws = wb.active
    if ws is None:
        ws = wb.create_sheet()
    ws.title = "Query Results"

    header_font = Font(bold=True)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    data_alignment = Alignment(vertical="center", wrap_text=True)
    centered_alignment = Alignment(
        horizontal="center", vertical="center", wrap_text=True
    )

    ws.append([header.replace("_", " ").title() for header in headers])
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.alignment = header_alignment

    for row in rows:
        ws.append([row.get(key, "") for key in headers])

    last_row = max(1, len(rows) + 1)
    last_col = len(headers)
    table_ref = f"A1:{get_column_letter(last_col)}{last_row}"
    table = Table(displayName="QueryResults", ref=table_ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleLight1",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(table)

    ws.column_dimensions["A"].width = 50
    ws.column_dimensions["B"].width = 50
    ws.column_dimensions["C"].width = 30
    ws.column_dimensions["D"].auto_size = True
    ws.column_dimensions["E"].width = 30
    ws.column_dimensions["F"].width = 80

    # for col_idx in range(1, len(headers) + 1):
    #     ws.column_dimensions[get_column_letter(col_idx)].width = 24

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            if cell.column in (4, 5):
                cell.alignment = centered_alignment
            else:
                cell.alignment = data_alignment

    summary = wb.create_sheet("Summary")
    summary_headers_query = ["query", "matches", "files_with_matches"]
    summary_headers_file = ["relative_path", "matches", "queries_matched"]

    query_counts: dict[str, int] = defaultdict(int)
    query_files: dict[str, set[str]] = defaultdict(set)
    file_counts: dict[str, int] = defaultdict(int)
    file_queries: dict[str, set[str]] = defaultdict(set)

    for row in rows:
        query = str(row.get("query", ""))
        relative_path = str(row.get("relative_path", ""))
        query_counts[query] += 1
        query_files[query].add(relative_path)
        file_counts[relative_path] += 1
        file_queries[relative_path].add(query)

    summary.append(
        [header.replace("_", " ").title() for header in summary_headers_query]
    )
    for col_idx, _ in enumerate(summary_headers_query, start=1):
        cell = summary.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.alignment = header_alignment

    for query in sorted(query_counts):
        summary.append(
            [
                query,
                query_counts[query],
                len(query_files[query]),
            ]
        )

    query_table_last_row = max(1, len(query_counts) + 1)
    query_table_ref = f"A1:C{query_table_last_row}"
    query_table = Table(displayName="QuerySummary", ref=query_table_ref)
    query_table.tableStyleInfo = TableStyleInfo(
        name="TableStyleLight1",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    summary.add_table(query_table)

    file_table_start = query_table_last_row + 2
    summary.cell(
        row=file_table_start,
        column=1,
        value=summary_headers_file[0].replace("_", " ").title(),
    )
    summary.cell(
        row=file_table_start,
        column=2,
        value=summary_headers_file[1].replace("_", " ").title(),
    )
    summary.cell(
        row=file_table_start,
        column=3,
        value=summary_headers_file[2].replace("_", " ").title(),
    )
    for col_idx in range(1, 4):
        cell = summary.cell(row=file_table_start, column=col_idx)
        cell.font = header_font
        cell.alignment = header_alignment

    for relative_path in sorted(file_counts):
        summary.append(
            [
                relative_path,
                file_counts[relative_path],
                len(file_queries[relative_path]),
            ]
        )

    file_table_last_row = file_table_start + len(file_counts)
    file_table_ref = f"A{file_table_start}:C{file_table_last_row}"
    file_table = Table(displayName="FileSummary", ref=file_table_ref)
    file_table.tableStyleInfo = TableStyleInfo(
        name="TableStyleLight1",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    summary.add_table(file_table)

    summary.column_dimensions["A"].width = 60
    summary.column_dimensions["B"].width = 18
    summary.column_dimensions["C"].width = 22
    if query_table_last_row >= 2:
        for row in summary.iter_rows(min_row=2, max_row=query_table_last_row):
            for cell in row:
                cell.alignment = data_alignment
    if file_table_last_row > file_table_start:
        for row in summary.iter_rows(
            min_row=file_table_start + 1, max_row=file_table_last_row
        ):
            for cell in row:
                cell.alignment = data_alignment

    xlsx_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(xlsx_path)


@sdk.script
def main(ctx: sdk.ScriptContext, api: sdk.ScriptAPI) -> None:
    recursive_field: sdk.BoolField = {
        "id": "recursive",
        "type": "bool",
        "label": "Recursive",
        "default": False,
    }

    payload = api.request_input_json(
        "Enter search query",
        id="query_pdf_options",
        fields=[
            *([recursive_field] if ctx.target_kind == "directory" else []),
            {
                "id": "case_sensitive",
                "type": "bool",
                "label": "Case sensitive",
                "default": False,
            },
        ],
        text_input_style="multiline",
        payload_type=QueryOptions,
    )

    raw_query = payload["value"]
    recursive = payload.get("recursive", False)
    case_sensitive = payload["case_sensitive"]

    if not raw_query:
        api.emit_result(
            {
                "_status": "warn",
                "_title": "Warning",
                "Info": "Search query is required. Query not performed.",
            }
        )
        return

    context_chars = 80
    context_response = api.request_input(
        "Context characters around each match",
        default=str(context_chars),
        id="query_pdf_context",
    )
    if context_response:
        try:
            context_chars = int(context_response)
        except ValueError:
            api.emit_result(
                {
                    "_status": "warn",
                    "_title": "Warning",
                    "Info": "Context characters must be a number.",
                }
            )
            return
        if context_chars <= 0:
            api.emit_result(
                {
                    "_status": "warn",
                    "_title": "Warning",
                    "Info": "Context characters must be greater than zero.",
                }
            )
            return

    target_path = ctx.target_path
    pdf_files = collect_files(
        target_path,
        "*.pdf",
        recursive=recursive,
        check_cancel=api.check_cancel,
    )
    root_path = target_path if ctx.target_kind == "directory" else target_path.parent
    total_files = len(pdf_files)
    if not total_files:
        api.emit_result(
            {
                "_status": "warn",
                "_title": "Warning",
                "Info": "No PDF files found to search.",
            }
        )
        return
    queries, invalid = _parse_queries(
        raw_query,
        case_sensitive=case_sensitive,
        api=api,
    )
    if invalid:
        api.emit_result(
            {
                "_status": "error",
                "_title": "Invalid Regex",
                "Info": "One or more regex patterns are invalid.",
                "Invalid": "; ".join(invalid),
            }
        )
        api.exit(code=1)
        return
    if not queries:
        api.emit_result(
            {
                "_status": "warn",
                "_title": "Warning",
                "Info": "No valid queries were provided.",
            }
        )
        return

    api.log(
        "info",
        (
            f"Queries={len(queries)} | recursive={recursive} "
            f"| case_sensitive={case_sensitive} | PDFs found={total_files}"
        ),
    )

    matches: list[dict[str, object]] = []
    files_with_matches: set[Path] = set()

    for index, pdf_path in enumerate(pdf_files, start=1):
        api.check_cancel()
        api.progress(current=index, total=total_files or 1, unit="files")
        try:
            pdf_matches = _process_pdf(
                pdf_path,
                root_path,
                queries,
                context_chars,
                api,
                check_cancel=api.check_cancel,
            )
        except Exception as exc:  # noqa: BLE001
            api.log("warn", f"Failed to process '{pdf_path}': {exc}")
            continue

        if pdf_matches:
            matches.extend(pdf_matches)
            files_with_matches.add(pdf_path)

    xlsx_path: Path | None = None
    if matches:
        stem = target_path.stem if ctx.target_kind == "file" else root_path.name
        output_dir = (
            target_path.parent if ctx.target_kind == "file" else root_path.parent
        )
        xlsx_path = output_dir / f"{stem}_query_results.xlsx"
        _write_xlsx(xlsx_path, matches)

    api.emit_result(
        {
            "_title": "PDF Query Results",
            "Queries": len(queries),
            "Files Searched": total_files,
            "Files With Matches": len(files_with_matches),
            "XLSX Path": str(xlsx_path) if xlsx_path else None,
        }
    )


if __name__ == "__main__":
    main()
