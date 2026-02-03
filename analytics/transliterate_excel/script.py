from __future__ import annotations

from typing import Any

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.utils import range_boundaries
from unidecode import unidecode

from ferp.fscp.scripts import sdk


def _normalize_range(value: str) -> str:
    return value.replace(" ", "").upper()


def _coerce_text(value: Any) -> str | None:
    if value is None:
        return None
    if isinstance(value, str):
        return value
    return None


@sdk.script
def main(ctx: sdk.ScriptContext, api: sdk.ScriptAPI) -> None:
    target_path = ctx.target_path

    range_value = api.request_input(
        "Enter cell range (e.g., A1:C21)",
        id="analytics_transliterate_excel_range",
    )
    range_value = _normalize_range(range_value or "")
    if not range_value:
        api.emit_result(
            {
                "_status": "error",
                "_title": "Error: Missing Range",
                "Info": "Enter a cell range like A1:C21.",
            }
        )
        return

    try:
        min_col, min_row, max_col, max_row = range_boundaries(range_value)
    except ValueError:
        api.emit_result(
            {
                "_status": "error",
                "_title": "Error: Invalid Range",
                "Info": "Enter a valid cell range like A1:C21.",
            }
        )
        return

    api.log("info", f"Loading workbook: {target_path.name}")
    try:
        workbook = load_workbook(filename=target_path)
    except Exception as exc:
        api.emit_result(
            {
                "_status": "error",
                "_title": "Error: Failed to Load Workbook",
                "Info": str(exc),
            }
        )
        return

    sheet = workbook.active
    if sheet is None:
        api.emit_result(
            {
                "_status": "error",
                "_title": "Error: Missing Worksheet",
                "Info": "Workbook does not contain an active worksheet.",
            }
        )
        return
    updated_cells = 0
    skipped_cells = 0

    for row in sheet.iter_rows(
        min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col
    ):
        for cell in row:
            if isinstance(cell, MergedCell):
                skipped_cells += 1
                continue
            text_value = _coerce_text(cell.value)
            if text_value is None:
                skipped_cells += 1
                continue
            converted = unidecode(text_value)
            if converted != text_value:
                cell.value = converted
                updated_cells += 1
            else:
                skipped_cells += 1

    try:
        workbook.save(target_path)
    except PermissionError:
        api.emit_result(
            {
                "_status": "error",
                "_title": "Error: File In Use",
                "Info": "Close the file in other programs and try again.",
            }
        )
        return
    except Exception as exc:
        api.emit_result(
            {
                "_status": "error",
                "_title": "Error: Failed to Save Workbook",
                "Info": str(exc),
            }
        )
        return

    api.emit_result(
        {
            "_status": "ok",
            "_title": "Transliteration Complete",
            "Updated cells": str(updated_cells),
            "Skipped cells": str(skipped_cells),
            "Range": range_value,
            "File": target_path.name,
        }
    )


if __name__ == "__main__":
    main()
